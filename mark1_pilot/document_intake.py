from __future__ import annotations

import base64
import io
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader


def _decode_bytes(content_base64: str) -> bytes:
    text = str(content_base64 or "").strip()
    if "," in text and text.lower().startswith("data:"):
        text = text.split(",", 1)[1]
    return base64.b64decode(text)


def _extract_text_from_pdf(data: bytes) -> tuple[str, dict[str, Any]]:
    reader = PdfReader(io.BytesIO(data))
    pages: list[str] = []
    for page in reader.pages[:5]:
        try:
            pages.append(page.extract_text() or "")
        except Exception:
            continue
    return "\n".join(pages).strip(), {"page_count": len(reader.pages)}


def _extract_text_from_workbook(data: bytes) -> tuple[str, dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet_summaries: list[dict[str, Any]] = []
    lines: list[str] = []
    for sheet in workbook.worksheets[:3]:
        preview_rows: list[list[str]] = []
        for row in sheet.iter_rows(max_row=8, max_col=8, values_only=True):
            values = [str(value).strip() for value in row if value not in {None, ""}]
            if values:
                preview_rows.append(values)
                lines.append(" | ".join(values))
        sheet_summaries.append({"sheet": sheet.title, "preview_rows": preview_rows[:5]})
    return "\n".join(lines).strip(), {"sheet_count": len(workbook.worksheets), "sheets": sheet_summaries}


def _extract_text_from_text(data: bytes) -> str:
    for encoding in ("utf-8", "utf-16", "latin-1"):
        try:
            return data.decode(encoding)
        except Exception:
            continue
    return data.decode("utf-8", errors="ignore")


def extract_document_text(filename: str, content_base64: str) -> tuple[str, dict[str, Any]]:
    suffix = Path(str(filename or "").strip()).suffix.lower()
    data = _decode_bytes(content_base64)

    if suffix == ".pdf":
        return _extract_text_from_pdf(data)
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_text_from_workbook(data)
    if suffix in {".json"}:
        try:
            payload = json.loads(_extract_text_from_text(data))
            return json.dumps(payload, indent=2), {"json": True}
        except Exception:
            return _extract_text_from_text(data), {"json": False}
    return _extract_text_from_text(data), {}


def _pick_document_type(filename: str, text: str) -> tuple[str, str]:
    lowered = f"{filename}\n{text}".lower()
    if re.search(r"(invoice|payment|remittance|overdue|bank|cash receive)", lowered):
        return "cash_watch", "Cash Watch"
    if re.search(r"(defect|complaint|reject|ncr|capa|quality)", lowered):
        return "quality_closeout", "Quality Closeout"
    if re.search(r"(stock|inventory|warehouse|on hand|reorder)", lowered):
        return "inventory_pulse", "Inventory Pulse"
    if re.search(r"(grn|goods received|received qty|packing list|batch|received_at)", lowered):
        return "receiving_control", "Receiving Control"
    if re.search(r"(supplier|eta|shipment|customs|po|pi)", lowered):
        return "supplier_watch", "Supplier Watch"
    return "action_os", "Action OS"


def _extract_first(pattern: str, text: str) -> str:
    match = re.search(pattern, text, flags=re.I)
    return match.group(1).strip() if match else ""


def analyze_document(filename: str, content_base64: str) -> dict[str, Any]:
    text, meta = extract_document_text(filename, content_base64)
    preview = re.sub(r"\s+", " ", text).strip()[:1200]
    document_type, recommended_module = _pick_document_type(filename, text)

    fields = {
      "supplier": _extract_first(r"(?:supplier|vendor)[:\s]+([-A-Za-z0-9 .,&/]+)", text),
      "invoice_no": _extract_first(r"(?:invoice(?: no| number)?|inv(?: no)?)[:#\s]+([-A-Za-z0-9/]+)", text),
      "po_or_pi": _extract_first(r"(?:po|p/o|purchase order|pi|proforma invoice)[:#\s]+([-A-Za-z0-9/]+)", text),
      "grn_or_batch": _extract_first(r"(?:grn|batch)[:#\s]+([-A-Za-z0-9/]+)", text),
      "item_code": _extract_first(r"(?:item code|material code|code)[:#\s]+([-A-Za-z0-9/]+)", text),
    }
    fields = {key: value for key, value in fields.items() if value}

    summary_map = {
      "cash_watch": "This looks like a finance or payment document. Route it into Cash Watch for invoice, overdue, or remittance follow-up.",
      "quality_closeout": "This looks like a quality or complaint document. Route it into Quality Closeout so it becomes an incident with owner and closure discipline.",
      "inventory_pulse": "This looks like a stock or warehouse document. Route it into Inventory Pulse to surface stock pressure and reorder action.",
      "receiving_control": "This looks like an inbound or GRN-style document. Route it into Receiving Control to track receipt, variance, and release status.",
      "supplier_watch": "This looks like a supplier or shipment document. Route it into Supplier Watch to monitor ETA, customs, and supplier risk.",
      "action_os": "This document should first land in Action OS so the team can decide the next owner, due date, and system route.",
    }

    return {
        "filename": filename,
        "document_type": document_type,
        "recommended_module": recommended_module,
        "summary": summary_map[document_type],
        "preview": preview,
        "extracted_fields": fields,
        "meta": meta,
    }
